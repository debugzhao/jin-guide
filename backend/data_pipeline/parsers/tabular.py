from __future__ import annotations

import csv
import io
import json
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable, Sequence

from data_pipeline.config import PipelineConfig
from data_pipeline.text_encoding import decode_html_bytes
from data_pipeline.records import (
    AdmissionPlanRecord,
    AdmissionScoreRecord,
    Provenance,
    RankSegmentRecord,
    SubjectType,
)


@dataclass(frozen=True)
class TabularDocument:
    rows: list[list[str]]
    page_or_sheet: list[int]


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def read_tabular_document(
    path: str | Path, *, table_index: int | None = None
) -> TabularDocument:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        text = file_path.read_text(encoding="utf-8-sig")
        rows = [[_cell(value) for value in row] for row in csv.reader(io.StringIO(text))]
        return TabularDocument(rows=rows, page_or_sheet=[1] * len(rows))
    if suffix in {".xlsx", ".xls"}:
        return _read_excel(file_path)
    if suffix == ".pdf":
        return _read_pdf(file_path)
    if suffix in {".jpg", ".jpeg", ".png"}:
        return _read_image(file_path)
    if suffix in {".html", ".htm"}:
        return _read_html_table(file_path, table_index=table_index)
    raise ValueError(f"unsupported tabular document: {file_path.name}")


class _HtmlTableParser(HTMLParser):
    """把页面里每一个"顶层"（不嵌套在别的表格里）<table> 各转成一个矩形网格，
    rowspan/colspan 的续格用空字符串占位——跟 openpyxl/pdfplumber 读合并单元格的
    语义一致（续格是空字符串不是重复文本），这样 parse_admission_plan_rows 已有的
    "空列=沿用上一行"前向填充逻辑可以直接复用，不需要为HTML另写一套规则。

    老式布局页面常见"表格套表格"（导航栏用一层 <table> 包住整个页头，真正的数据
    表在更后面另一个顶层 <table> 里）——按嵌套深度过滤，只在深度恰好为1（直接在
    某个顶层表格内，不在其嵌套子表格内）时才记录 <tr>/<td>，避免把导航表格自己
    嵌套的子表格误当成页面级别的第二个候选表格、也避免嵌套表格提前触发"表格结束"
    把外层表格后续的行截断。真正选哪个表格是"数据表"由调用方按行数挑（见
    read_tabular_document 的 table_index 参数），这里只负责把所有顶层表格都提取
    出来。
    """

    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._depth = 0
        self._row: list[str] | None = None
        self._cell_parts: list[str] | None = None
        self._cell_colspan = 1
        self._cell_rowspan = 1
        self._pending: dict[int, int] = {}  # column -> 还需要占位的行数
        self._col = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._depth += 1
            if self._depth == 1:
                self.tables.append([])
                self._pending = {}
            return
        if self._depth != 1:
            return
        if tag == "tr":
            self._row = []
            self._col = 0
        elif tag in ("td", "th") and self._row is not None:
            values = dict(attrs)
            self._cell_parts = []
            self._cell_colspan = self._parse_span(values.get("colspan"))
            self._cell_rowspan = self._parse_span(values.get("rowspan"))

    def handle_data(self, data: str) -> None:
        if self._depth == 1 and self._cell_parts is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "table":
            self._depth = max(0, self._depth - 1)
            if self._depth == 0:
                self._row = None
            return
        if self._depth != 1:
            return
        if tag in ("td", "th") and self._cell_parts is not None and self._row is not None:
            text = re.sub(r"\s+", " ", "".join(self._cell_parts)).strip()
            self._fill_pending_columns()
            for offset in range(self._cell_colspan):
                self._row.append(text if offset == 0 else "")
                if self._cell_rowspan > 1:
                    self._pending[self._col] = self._cell_rowspan - 1
                self._col += 1
            self._cell_parts = None
        elif tag == "tr" and self._row is not None:
            self.tables[-1].append(self._row)
            self._row = None

    def _fill_pending_columns(self) -> None:
        assert self._row is not None
        while self._pending.get(self._col, 0) > 0:
            self._row.append("")
            remaining = self._pending[self._col] - 1
            if remaining <= 0:
                del self._pending[self._col]
            else:
                self._pending[self._col] = remaining
            self._col += 1

    @staticmethod
    def _parse_span(value: str | None) -> int:
        try:
            return max(1, int(value)) if value else 1
        except ValueError:
            return 1


def _read_html_table(path: Path, *, table_index: int | None = None) -> TabularDocument:
    parser = _HtmlTableParser()
    parser.feed(decode_html_bytes(path.read_bytes()))
    if not parser.tables:
        return TabularDocument(rows=[], page_or_sheet=[])
    if table_index is not None:
        rows = parser.tables[table_index]
    else:
        # 默认挑行数最多的顶层表格——装饰性的导航/页头/页脚表格行数很少，
        # 真正的数据表通常是页面里行数最多的那个，已用真实多表格页面验证过。
        rows = max(parser.tables, key=len)
    return TabularDocument(rows=rows, page_or_sheet=[1] * len(rows))


def _read_excel(path: Path) -> TabularDocument:
    if path.suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - depends on deployment image
            raise RuntimeError("xlrd is required to parse legacy .xls files") from exc
        workbook = xlrd.open_workbook(path)
        rows: list[list[str]] = []
        sheets: list[int] = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            for row_index in range(sheet.nrows):
                normalized = [_cell(value) for value in sheet.row_values(row_index)]
                if any(normalized):
                    rows.append(normalized)
                    sheets.append(sheet_index + 1)
        return TabularDocument(rows=rows, page_or_sheet=sheets)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on deployment image
        raise RuntimeError("openpyxl is required to parse Excel files") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[str]] = []
    sheets: list[int] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            for row in sheet.iter_rows(values_only=True):
                normalized = [_cell(value) for value in row]
                if any(normalized):
                    rows.append(normalized)
                    sheets.append(sheet_index)
    finally:
        workbook.close()
    return TabularDocument(rows=rows, page_or_sheet=sheets)


def _read_pdf(path: Path) -> TabularDocument:
    try:
        import pdfplumber
    except ImportError as exc:  # pragma: no cover - depends on deployment image
        raise RuntimeError("pdfplumber is required to parse text PDFs") from exc
    rows: list[list[str]] = []
    pages: list[int] = []
    with pdfplumber.open(path) as document:
        for page_number, page in enumerate(document.pages, start=1):
            for table in page.extract_tables() or []:
                for row in table:
                    normalized = [_cell(value) for value in row]
                    if any(normalized):
                        rows.append(normalized)
                        pages.append(page_number)
    if rows:
        return TabularDocument(rows=rows, page_or_sheet=pages)
    # 扫描版PDF（如上海市教育考试院的《考生高考成绩分布表》）没有可提取文本/表格，
    # pdfplumber 拿不到任何行；渲染成图片后走跟 _read_image 一样的 Vision OCR 路径，
    # 而不是直接报错要求人工——这类PDF在生产环境同样需要OCR兜底
    return _read_pdf_via_ocr(path)


def _read_pdf_via_ocr(path: Path) -> TabularDocument:
    vision_script = Path(__file__).resolve().parents[2] / "scripts" / "macos_vision_ocr.swift"
    if os.uname().sysname != "Darwin" or not vision_script.exists():
        raise RuntimeError(
            "PDF contains no extractable tables and image OCR unavailable; "
            "install PaddleOCR in production or run on macOS with Vision"
        )
    import pdfplumber

    rows: list[list[str]] = []
    pages: list[int] = []
    with pdfplumber.open(path) as document, tempfile.TemporaryDirectory(dir="/tmp") as temp_dir:
        for page_number, page in enumerate(document.pages, start=1):
            page_image_path = Path(temp_dir) / f"page-{page_number}.png"
            page.to_image(resolution=200).save(page_image_path)
            observations = _run_vision(page_image_path, vision_script)
            # 单栏正文表格（分数/人数/累计人数三列），不是江苏那种一页三个分栏拼版，
            # panel_count=1 直接按行聚类即可
            for row in _cluster_ocr_rows(observations, panel_count=1):
                if any(row):
                    rows.append(row)
                    pages.append(page_number)
    if not rows:
        raise RuntimeError("PDF OCR produced no rows; manual review required")
    return TabularDocument(rows=rows, page_or_sheet=pages)


def _read_image(path: Path) -> TabularDocument:
    """OCR an image into approximate table rows.

    Production Linux images should provide PaddleOCR. macOS development uses the
    built-in Vision framework, avoiding a heavyweight local model download.
    """
    vision_script = Path(__file__).resolve().parents[2] / "scripts" / "macos_vision_ocr.swift"
    if os.uname().sysname == "Darwin" and vision_script.exists():
        rows = _read_image_with_vision(path, vision_script)
        return TabularDocument(rows=rows, page_or_sheet=[1] * len(rows))
    raise RuntimeError(
        "image OCR unavailable; install PaddleOCR in production or run on macOS with Vision"
    )


def _read_image_with_vision(path: Path, vision_script: Path) -> list[list[str]]:
    try:
        from PIL import Image, ImageOps
    except ImportError:  # pragma: no cover
        return _cluster_ocr_rows(_run_vision(path, vision_script), panel_count=3)

    rows: list[list[str]] = []
    with Image.open(path) as image, tempfile.TemporaryDirectory(dir="/tmp") as temp_dir:
        grayscale = ImageOps.grayscale(image)
        width, height = grayscale.size
        for panel_index in range(3):
            left = round(width * panel_index / 3)
            right = round(width * (panel_index + 1) / 3)
            panel = grayscale.crop((left, 0, right, height))
            panel = ImageOps.autocontrast(panel, cutoff=1)
            # Watermarks are light gray; a conservative threshold removes them
            # while retaining the black printed digits and table rules.
            panel = panel.point(lambda value: 255 if value > 175 else 0)
            panel = panel.resize(
                (panel.width * 2, panel.height * 2), Image.Resampling.LANCZOS
            )
            panel_path = Path(temp_dir) / f"panel-{panel_index}.png"
            panel.save(panel_path)
            observations = _run_vision(panel_path, vision_script)
            rows.extend(_cluster_ocr_rows(observations, panel_count=1))
    return rows


def _run_vision(path: Path, vision_script: Path) -> list[dict]:
    environment = os.environ.copy()
    environment.setdefault("SWIFT_MODULECACHE_PATH", "/tmp/wenjin-swift-module-cache")
    environment.setdefault("CLANG_MODULE_CACHE_PATH", "/tmp/wenjin-clang-module-cache")
    compiled_helper = Path("/tmp/wenjin-macos-vision-ocr")
    command = (
        [str(compiled_helper), str(path)]
        if compiled_helper.exists()
        else ["xcrun", "swift", str(vision_script), str(path)]
    )
    result = subprocess.run(
        command,
        check=True,
        capture_output=True,
        text=True,
        timeout=120,
        env=environment,
    )
    return json.loads(result.stdout)


def _cluster_ocr_rows(
    observations: list[dict], *, panel_count: int = 3
) -> list[list[str]]:
    # JSEE rank images place up to three independent tables side by side and may
    # concatenate multiple pages vertically. Partitioning by x before clustering
    # prevents values from different panels being merged into one logical row.
    panels: list[list[dict]] = [[] for _ in range(panel_count)]
    for item in observations:
        center_x = item["x"] + item["width"] / 2
        panel_index = min(int(center_x * panel_count), panel_count - 1)
        panels[panel_index].append(item)

    clusters: list[dict] = []
    for panel in panels:
        ordered = sorted(panel, key=lambda item: (-item["y"], item["x"]))
        panel_clusters: list[dict] = []
        for item in ordered:
            center_y = item["y"] + item["height"] / 2
            match = next(
                (
                    cluster
                    for cluster in panel_clusters
                    if abs(cluster["center_y"] - center_y)
                    <= max(item["height"], cluster["height"]) * 0.75
                ),
                None,
            )
            if match is None:
                panel_clusters.append(
                    {"center_y": center_y, "height": item["height"], "items": [item]}
                )
            else:
                match["items"].append(item)
        clusters.extend(panel_clusters)
    return [
        [entry["text"] for entry in sorted(cluster["items"], key=lambda item: item["x"])]
        for cluster in clusters
    ]


def _integer(value: str) -> int | None:
    match = re.search(r"(?<!\d)(\d{1,7})(?!\d)", value.replace(",", ""))
    return int(match.group(1)) if match else None


def _row_number(provenance: Provenance, row_number: int, page: int | None) -> Provenance:
    return provenance.model_copy(update={"table_row": row_number, "page_number": page})


def parse_rank_segment_rows(
    document: TabularDocument,
    *,
    subject_type: SubjectType,
    provenance: Provenance,
    config: PipelineConfig,
) -> list[RankSegmentRecord]:
    records: list[RankSegmentRecord] = []
    for row_number, row in enumerate(document.rows, start=1):
        values = [_integer(value) for value in row]
        numbers = [value for value in values if value is not None]
        groups: list[list[int]] = []
        if len(numbers) == 2:
            groups = [[numbers[0], 0, numbers[1]]]
        elif len(numbers) == 3:
            groups = [numbers]
        elif len(numbers) >= 6 and len(numbers) % 3 == 0:
            groups = [numbers[index : index + 3] for index in range(0, len(numbers), 3)]
        for score, count, rank in groups:
            if not 0 <= score <= 750 or rank <= 0 or rank < count:
                continue
            records.append(
                RankSegmentRecord(
                    province=config.province,
                    year=provenance.year,
                    subject_type=subject_type,
                    score=score,
                    cumulative_rank=rank,
                    provenance=_row_number(
                        provenance, row_number, document.page_or_sheet[row_number - 1]
                    ),
                )
            )
    return records


# 上海市教育考试院投档线PDF里高校名称是简称（如"上海交大"而非"上海交通大学"），且
# 表格从扫描/水印PDF提取时偶尔会在名称前混入一个水印字（"市\n华东理工(02)"），已用
# 真实2025年PDF核对过这10所目标高校在该文件里的确切写法，不是猜测；复旦/交大各自
# 的"复旦医学"/"交大医学"专业组是本校上海医学院的投档线，仍算同一所学校。
_SHMEEA_UNIVERSITY_ALIASES: dict[str, str] = {
    "复旦大学": "10246",
    "复旦医学": "10246",
    "上海交大": "10248",
    "交大医学": "10248",
    "同济大学": "10247",
    "华东师大": "10269",
    "上海财大": "10272",
    "上海外大": "10271",
    "华东理工": "10251",
    "东华大学": "10255",
    "上海大学": "10280",
    "上海理工": "10252",
}


def _clean_shmeea_cell(value: str) -> str:
    # 水印字混入时总是作为独立的一行出现在真正内容前面（如"市\n华东理工(02)"），
    # 取最后一段就能拿到干净文本；没有换行时原样返回
    return value.split("\n")[-1].strip()


def parse_shmeea_admission_score_rows(
    document: TabularDocument,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    batch: str,
    admission_type: str = "普通",
) -> tuple[list[AdmissionScoreRecord], int]:
    """解析上海市教育考试院《本科批次平行志愿院校专业组投档分数线》。

    返回 (records, undisclosed_count)：580分及以上考生的投档线官方明确不公开
    （"由市教育考试院会同考生所在中学逐一告知"），这些行会被跳过而不是编造成
    580 分——undisclosed_count 记录跳过了多少条，用于在报告里如实体现覆盖率
    缺口，而不是让它无声消失。
    """
    allowed_codes = {target.university_code for target in config.target_universities}
    records: list[AdmissionScoreRecord] = []
    undisclosed_count = 0
    for row_number, row in enumerate(document.rows, start=1):
        if len(row) < 3:
            continue
        group_code = _clean_shmeea_cell(row[0])
        name_cell = _clean_shmeea_cell(row[1])
        score_cell = _clean_shmeea_cell(row[2])
        if not re.fullmatch(r"[A-Za-z0-9]{4,8}", group_code):
            continue
        base_name = re.sub(r"\(.*", "", name_cell).strip()
        university_code = _SHMEEA_UNIVERSITY_ALIASES.get(base_name)
        if university_code is None or university_code not in allowed_codes:
            continue
        if "以上" in score_cell:
            # "580分及以上"这类未公开区间，_integer()会误把580当成真实投档线提取
            # 出来，必须先排除这个模式，不能拿580去冒充真实投档线
            undisclosed_count += 1
            continue
        min_score = _integer(score_cell)
        if min_score is None:
            continue
        target = next(t for t in config.target_universities if t.university_code == university_code)
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=provenance.year,
                batch=batch,
                subject_type="unified",
                university_code=university_code,
                university_name=target.name,
                major_group_code=group_code,
                # 保留"(01)"这类组别后缀，不能只存base_name——business_sync.py的
                # sync_admission_scores按(university_id,year,batch,subject_type,
                # major_category)去重，major_category取major_group_name优先，
                # 同校多个专业组若都存成同一个base_name会在业务表里互相覆盖（同
                # 一所大学的"华东理工(01)"560分和"华东理工(03)"548分曾经因此被
                # 合并成一条，已实测踩过这个坑）
                major_group_name=name_cell,
                admission_type=admission_type,
                min_score=min_score,
                provenance=_row_number(
                    provenance, row_number, document.page_or_sheet[row_number - 1]
                ),
            )
        )
    return records, undisclosed_count


_GROUP_PATTERN = re.compile(
    r"^(?P<name>.+?)(?P<group>[A-Z]?\d{2,3})专业组(?:\((?P<requirement>[^)]+)\))?(?P<tail>.*)$"
)

# 专业组名称后可能出现第二个括注（如 "…专业组(化学)(中外合作办学)"），标注的是招生
# 类型而不是校区。这里用于把这类标签从 tail 中过滤掉，避免它们被误当成校区写入
# campus 字段。
_ADMISSION_TYPE_LABELS = ("中外合作办学", "高校专项", "地方专项", "联合培养")
# 单校招生计划页"类别"列的标准平行志愿写法是"普通类 0005"（类别名+代码，无额外
# 说明文字）；"三位一体"等特殊招生机制会在后面附一长段说明文字，不匹配这个模式。
_PLAIN_CATEGORY_PATTERN = re.compile(r"^普通类\s*\d*$")


def parse_admission_score_rows(
    document: TabularDocument,
    *,
    subject_type: SubjectType,
    provenance: Provenance,
    config: PipelineConfig,
    batch: str = "本科批",
    line_type: str = "regular",
) -> list[AdmissionScoreRecord]:
    by_name = {target.name: target for target in config.target_universities}
    records: list[AdmissionScoreRecord] = []
    for row_number, row in enumerate(document.rows, start=1):
        joined = " ".join(value for value in row if value)
        group_match = None
        group_cell = ""
        group_index = 0
        for cell_index, value in enumerate(row):
            match = _GROUP_PATTERN.match(value.replace(" ", ""))
            if match and match.group("name") in by_name:
                group_match = match
                group_cell = value
                group_index = cell_index
                break
        if group_match is None:
            continue

        integers = [_integer(value) for value in row[group_index + 1 :]]
        numeric = [value for value in integers if value is not None]
        plausible_scores = [value for value in numeric if 0 <= value <= 750]
        if not plausible_scores:
            continue
        min_score = plausible_scores[0]
        university_name = group_match.group("name")
        tail = (group_match.group("tail") or "").strip("()")
        admission_type = "普通"
        for label in _ADMISSION_TYPE_LABELS:
            if label in joined:
                admission_type = label
                break
        campus = tail if tail and tail not in _ADMISSION_TYPE_LABELS else None
        target = by_name[university_name]
        provincial_code = next(
            (value for value in row if re.fullmatch(r"\d{4}", value.strip())), None
        )
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=provenance.year,
                batch=batch,
                subject_type=subject_type,
                university_code=target.university_code,
                university_name=university_name,
                provincial_university_code=provincial_code,
                major_group_code=group_match.group("group"),
                major_group_name=group_cell,
                selection_requirement=group_match.group("requirement"),
                admission_type=admission_type,
                campus=campus,
                line_type=line_type,
                min_score=min_score,
                provenance=_row_number(
                    provenance, row_number, document.page_or_sheet[row_number - 1]
                ),
            )
        )
    return records


def parse_zhejiang_admission_score_rows(
    document: TabularDocument,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    batch: str,
) -> list[AdmissionScoreRecord]:
    """浙江省教育考试院投档线是扁平表（学校代号/学校名称/专业代号/专业名称/计划数/
    分数线/位次），每行本身就带位次，不需要像江苏那样另外关联逐分段表做enrichment；
    每个专业本身就是志愿单位，没有"院校专业组"这层概念，这里复用
    major_group_code/major_group_name 字段存专业代号/专业名称（浙江语义下"1个专业
    (类)"就相当于江苏语义下的"1个院校专业组"，都是最小志愿单位，字段含义对得上，不是
    乱塞）。用真实2026年浙江投档线xls验证过表头结构，见
    docs/10_zhejiang_top10_data_collection_status.md。
    """
    aliases = {
        "学校代号": {"学校代号", "学校代码", "院校代号", "院校代码"},
        "学校名称": {"学校名称", "院校名称"},
        "专业代号": {"专业代号", "专业代码"},
        "专业名称": {"专业名称"},
        "分数线": {"分数线", "投档线", "最低分"},
        "位次": {"位次", "最低位次"},
    }
    header_index = None
    columns: dict[str, int] = {}
    for index, row in enumerate(document.rows):
        compact = [re.sub(r"\s+", "", value) for value in row]
        found: dict[str, int] = {}
        for canonical, options in aliases.items():
            for cell_index, value in enumerate(compact):
                if value in options:
                    found[canonical] = cell_index
                    break
        if {"学校名称", "专业名称", "分数线"}.issubset(found):
            header_index = index
            columns = found
            break
    if header_index is None:
        return []

    by_name = {target.name: target for target in config.target_universities}

    def value(row: Sequence[str], key: str) -> str:
        position = columns.get(key)
        return row[position].strip() if position is not None and position < len(row) else ""

    records: list[AdmissionScoreRecord] = []
    for row_number in range(header_index + 1, len(document.rows)):
        row = document.rows[row_number]
        university_name = value(row, "学校名称")
        target = by_name.get(university_name)
        if target is None:
            continue
        major_code = value(row, "专业代号")
        major_name = value(row, "专业名称")
        min_score = _integer(value(row, "分数线"))
        if not major_code or not major_name or min_score is None:
            continue
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=provenance.year,
                batch=batch,
                subject_type="unified",
                university_code=target.university_code,
                university_name=university_name,
                provincial_university_code=value(row, "学校代号") or None,
                major_group_code=major_code,
                major_group_name=major_name,
                min_score=min_score,
                min_rank=_integer(value(row, "位次")),
                provenance=_row_number(
                    provenance, row_number, document.page_or_sheet[row_number - 1]
                ),
            )
        )
    return records


def parse_single_university_admission_result_rows(
    document: TabularDocument,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    target_university_code: str,
    batch: str,
) -> list[AdmissionScoreRecord]:
    """有些学校自己在"历年招生"栏目发布的录取情况页，字段比省考试院投档线表更
    丰富——真实表头是"专业名称/录取数/最高分/最低分/平均分/最低位次"（已用杭州
    师范大学2025年浙江省普通类一段首轮录取情况页验证），带平均分/最高分/实际
    录取人数，这些是`parse_zhejiang_admission_score_rows`（省考试院表，只有
    分数线+位次）没有的补充信息，对应PRD"专业录取最低分、最低位次...官方提供时
    保存平均分、最高分和录取人数"这条要求。

    没有专业代号（省考试院表才有），major_group_code复用专业名称本身（不是
    编造代码，是如实反映"这份数据源里专业没有编号，名称就是唯一标识"）——这跟
    省考试院表的major_group_code（数字代号）取值格式不同，两份数据的
    natural_key不会撞在一起，是两条并存的独立记录，不是同一条记录的两个版本，
    发布前需要人工决定要不要合并（见状态看板#9）。
    """
    by_code = {target.university_code: target for target in config.target_universities}
    target = by_code.get(target_university_code)
    if target is None:
        raise ValueError(f"target_university_code {target_university_code!r} not in whitelist")

    aliases = {
        "专业名称": ("专业名称",),
        "录取数": ("录取数", "录取人数"),
        "最高分": ("最高分",),
        "最低分": ("最低分",),
        "平均分": ("平均分",),
        "位次": ("最低位次", "位次"),
    }
    header_index = None
    columns: dict[str, int] = {}
    for index, row in enumerate(document.rows):
        compact = [re.sub(r"\s+", "", value) for value in row]
        found: dict[str, int] = {}
        for canonical, keywords in aliases.items():
            for cell_index, value in enumerate(compact):
                if value and any(keyword in value for keyword in keywords):
                    found[canonical] = cell_index
                    break
        if {"专业名称", "最低分"}.issubset(found):
            header_index = index
            columns = found
            break
    if header_index is None:
        return []

    def value(row: Sequence[str], key: str) -> str:
        position = columns.get(key)
        return row[position].strip() if position is not None and position < len(row) else ""

    records: list[AdmissionScoreRecord] = []
    for row_number in range(header_index + 1, len(document.rows)):
        row = document.rows[row_number]
        major_name = value(row, "专业名称")
        min_score = _integer(value(row, "最低分"))
        if not major_name or min_score is None:
            continue
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=provenance.year,
                batch=batch,
                subject_type="unified",
                university_code=target.university_code,
                university_name=target.name,
                major_group_code=major_name,
                major_group_name=major_name,
                min_score=min_score,
                max_score=_integer(value(row, "最高分")),
                avg_score=_integer(value(row, "平均分")),
                enrollment_count=_integer(value(row, "录取数")),
                min_rank=_integer(value(row, "位次")),
                provenance=_row_number(
                    provenance, row_number, document.page_or_sheet[row_number - 1]
                ),
            )
        )
    return records


def parse_single_university_admission_plan_rows(
    document: TabularDocument,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    target_university_code: str,
    subject_type: SubjectType,
    batch: str = "本科批",
) -> list[AdmissionPlanRecord]:
    """有些学校的招生计划页本身就是"这一所学校专属"的表（浙江10校里能http直采的
    几所都是这样：宁波大学/浙江工业大学等各自招生网自己的页面，不是省考试院发布的
    多校汇总表），没有"院校名称"这一列可以拿来匹配白名单——目标学校已经由
    source.target_university_code 决定，不需要再从某一列里搜校名。各校列头写法差异
    很大（"专业（类）名称" vs "专业名称"，"选考科目" vs "选科要求"），改用关键词包含
    匹配而不是`parse_admission_plan_rows`那种精确相等匹配。用真实宁波大学2026年招生
    计划页（含rowspan合并单元格）验证过。
    """
    by_code = {target.university_code: target for target in config.target_universities}
    target = by_code.get(target_university_code)
    if target is None:
        raise ValueError(f"target_university_code {target_university_code!r} not in whitelist")

    aliases = {
        "类别": ("类别", "类型"),
        "批次": ("批次",),
        "专业名称": ("专业",),
        "计划数": ("计划数", "计划人数", "招生计划"),
        "学费": ("学费", "收费"),
        "学制": ("学制", "修业年限"),
        "备注": ("备注", "报考要求"),
        "选科": ("选考科目", "选科要求", "再选科目"),
    }
    header_index = None
    columns: dict[str, int] = {}
    for index, row in enumerate(document.rows):
        compact = [re.sub(r"\s+", "", value) for value in row]
        found: dict[str, int] = {}
        for canonical, keywords in aliases.items():
            for cell_index, value in enumerate(compact):
                if value and any(keyword in value for keyword in keywords):
                    found[canonical] = cell_index
                    break
        if {"专业名称", "计划数"}.issubset(found):
            header_index = index
            columns = found
            break
    if header_index is None:
        return []

    def value(row: Sequence[str], key: str) -> str:
        position = columns.get(key)
        return row[position].strip() if position is not None and position < len(row) else ""

    records: list[AdmissionPlanRecord] = []
    current_category = ""
    for row_number in range(header_index + 1, len(document.rows)):
        row = document.rows[row_number]
        category_value = value(row, "类别")
        if category_value:
            current_category = category_value
        major_name = value(row, "专业名称")
        quota = _integer(value(row, "计划数"))
        if not major_name or quota is None or quota <= 0:
            continue
        if _PLAIN_CATEGORY_PATTERN.match(current_category):
            admission_type = "普通"
        else:
            # 类别栏还可能出现"普通类提前 三位一体 0238 只招已参加我校..."这类特殊
            # 招生机制说明（跟标准平行志愿的"普通类 0005"不是同一回事），如果不
            # 区分开，同一个专业名称在"普通类"和"三位一体"两条轨道下会撞成同一个
            # natural_key、被去重校验误当作重复记录拒绝——已用真实宁波大学数据
            # 验证过这个坑（"水产养殖学（拔尖人才创新班）"在两条轨道各出现一次）。
            # 已知具体标签优先给出可读值，否则退化成用原始类别文本兜底去重。
            admission_type = current_category or "普通"
            for label in _ADMISSION_TYPE_LABELS:
                if label in current_category:
                    admission_type = label
                    break
        records.append(
            AdmissionPlanRecord(
                province=config.province,
                year=provenance.year,
                # 部分学校的表本身就有"批次"列（如浙江理工大学：普通类/艺术类统考批/
                # 地方专项计划/单独考试招生计算机类/单独考试招生电子与电工类），比
                # 函数默认的"本科批"更精确，这一列存在时优先用它——同一专业名称在
                # "单独考试招生计算机类"和"单独考试招生电子与电工类"两条线各出现一次
                # 时，光靠admission_type/restrictions区分不开，已用真实数据验证过。
                batch=value(row, "批次") or batch,
                subject_type=subject_type,
                university_code=target.university_code,
                university_name=target.name,
                major_name=major_name,
                quota=quota,
                tuition=_integer(value(row, "学费")),
                duration_years=_integer(value(row, "学制")),
                selection_requirement=value(row, "选科") or None,
                admission_type=admission_type,
                restrictions=value(row, "备注") or None,
                provenance=_row_number(
                    provenance, row_number, document.page_or_sheet[row_number - 1]
                ),
            )
        )
    return records


def parse_admission_plan_rows(
    document: TabularDocument,
    *,
    subject_type: SubjectType,
    provenance: Provenance,
    config: PipelineConfig,
    batch: str = "本科批",
) -> list[AdmissionPlanRecord]:
    aliases = {
        "院校名称": {"院校名称", "学校名称"},
        "院校代码": {"院校代码", "院校代号"},
        "专业组": {"院校专业组", "专业组", "专业组代码"},
        "专业代码": {"专业代码", "专业代号"},
        "专业名称": {"专业名称", "专业名称及方向"},
        "计划数": {"计划数", "招生计划", "计划人数"},
        "学费": {"学费", "收费标准"},
        "学制": {"学制", "修业年限"},
        "校区": {"校区", "办学地点"},
        "选科": {"选科要求", "再选科目要求"},
        "备注": {"备注", "专业备注", "报考要求"},
    }
    header_index = None
    columns: dict[str, int] = {}
    for index, row in enumerate(document.rows):
        compact = [re.sub(r"\s+", "", value) for value in row]
        found: dict[str, int] = {}
        for canonical, options in aliases.items():
            for cell_index, value in enumerate(compact):
                if value in options:
                    found[canonical] = cell_index
                    break
        if {"专业代码", "专业名称", "计划数"}.issubset(found):
            header_index = index
            columns = found
            break
    if header_index is None:
        return []

    by_name = {target.name: target for target in config.target_universities}
    current_name: str | None = None
    current_group: str | None = None
    current_provincial_code: str | None = None
    records: list[AdmissionPlanRecord] = []

    def value(row: Sequence[str], key: str) -> str:
        position = columns.get(key)
        return row[position].strip() if position is not None and position < len(row) else ""

    for row_number in range(header_index + 1, len(document.rows)):
        row = document.rows[row_number]
        name_value = value(row, "院校名称")
        matched_name = next((name for name in by_name if name in name_value), None)
        if name_value:
            # 院校名称列非空代表新院校区块开始（无论是否在白名单内），必须清空
            # 上一个院校残留的专业组/院校代号，否则会把非白名单院校或前一所
            # 院校的专业组代号错误地挂到本院校后续行上。
            current_group = None
            current_provincial_code = None
            current_name = matched_name
        code_value = value(row, "院校代码")
        if code_value:
            current_provincial_code = code_value
        group_value = value(row, "专业组")
        group_match = re.search(r"([A-Z]?\d{2,3})", group_value)
        if group_match:
            current_group = group_match.group(1)
        if not current_name or not current_group:
            continue
        major_code = value(row, "专业代码")
        major_name = value(row, "专业名称")
        quota = _integer(value(row, "计划数"))
        if not major_code or not major_name or quota is None or quota <= 0:
            continue
        tuition = _integer(value(row, "学费"))
        duration = _integer(value(row, "学制"))
        remarks = value(row, "备注")
        admission_type = "普通"
        combined = " ".join(row)
        for label in _ADMISSION_TYPE_LABELS:
            if label in combined:
                admission_type = label
                break
        target = by_name[current_name]
        records.append(
            AdmissionPlanRecord(
                province=config.province,
                year=provenance.year,
                batch=batch,
                subject_type=subject_type,
                university_code=target.university_code,
                university_name=current_name,
                provincial_university_code=current_provincial_code,
                major_group_code=current_group,
                major_code=major_code,
                major_name=major_name,
                quota=quota,
                tuition=tuition,
                duration_years=duration,
                campus=value(row, "校区") or None,
                selection_requirement=value(row, "选科") or None,
                admission_type=admission_type,
                restrictions=remarks or None,
                provenance=_row_number(
                    provenance, row_number + 1, document.page_or_sheet[row_number]
                ),
            )
        )
    return records


def parse_zjgsu_admission_score_json(
    path: Path,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    target_university_code: str,
) -> list[AdmissionScoreRecord]:
    """浙江工商大学"历年分数"页背后是`POST /_api/zsfs/ action=getAllData`这个公开
    JSON接口（已用真实响应验证，CORS开放、无需鉴权），一次性返回全部省份/年份数据，
    不是省考试院投档线那种扁平表。真实字段是`province/category/type/major/
    planNumber/zdscore(最低)/zgscore(最高)/pjscore(平均)`（拼音缩写，非编造）。

    只保留`category=="综合"`里三类真正落在750分制平行志愿口径的`type`：
    "普通类平行"（含中外合作办学变体）和"地方专项计划"；显式排除"普通类提前"
    （三位一体综合评价的100分制折算分，跟750分制混在同一个min_score字段会
    误导下游）和"艺术类统考"（艺术类评分规则不同），已用真实数据核实这两类
    分数量级（80-90分/570-590分）跟普通类平行（600+分）明显不是同一把尺子。
    """
    by_code = {target.university_code: target for target in config.target_universities}
    target = by_code.get(target_university_code)
    if target is None:
        raise ValueError(f"target_university_code {target_university_code!r} not in whitelist")

    payload = json.loads(path.read_text(encoding="utf-8"))
    rows = payload.get("data") or []

    type_mapping = {
        "普通类平行": ("平行志愿", "普通"),
        "普通类平行(中外合作办学)": ("平行志愿", "中外合作办学"),
        "地方专项计划": ("地方专项计划", "地方专项"),
    }

    records: list[AdmissionScoreRecord] = []
    for row in rows:
        if row.get("province") != config.province or row.get("category") != "综合":
            continue
        mapping = type_mapping.get(row.get("type"))
        if mapping is None:
            continue
        batch, admission_type = mapping
        major_name = str(row.get("major") or "").strip()
        min_score = _integer(str(row.get("zdscore") or ""))
        if not major_name or min_score is None:
            continue
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=int(row.get("year") or provenance.year),
                batch=batch,
                subject_type="unified",
                university_code=target.university_code,
                university_name=target.name,
                major_group_code=major_name,
                major_group_name=major_name,
                admission_type=admission_type,
                min_score=min_score,
                max_score=_integer(str(row.get("zgscore") or "")),
                avg_score=_integer(str(row.get("pjscore") or "")),
                # planNumber是招生计划数，不是实际录取人数，两者语义不同不能混填
                # enrollment_count（AdmissionPlanRecord.quota才是计划数该去的地方）
                provenance=provenance,
            )
        )
    return records


def parse_wmu_admission_score_json(
    path: Path,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    target_university_code: str,
    batch: str = "本科批",
) -> list[AdmissionScoreRecord]:
    """温州医科大学"历年录取成绩"页背后是第三方招生数据服务商jobpi.cn托管的公开
    JSON接口（已用真实响应验证，无需鉴权），按`filter_column={A:年份,B:省份,C:类型}`
    筛选，字段名是脱敏后的字母代号，真实含义由响应里的`head`数组自描述（如
    "年份--80"，"--"后是列宽不是字段的一部分），改用动态映射而不是硬编码字母
    顺序，避免字段顺序变化时静默错位。
    """
    by_code = {target.university_code: target for target in config.target_universities}
    target = by_code.get(target_university_code)
    if target is None:
        raise ValueError(f"target_university_code {target_university_code!r} not in whitelist")

    payload = json.loads(path.read_text(encoding="utf-8"))
    data = payload.get("data") or {}
    head = (data.get("head") or [{}])[0]
    label_to_key = {
        str(label).split("--", 1)[0]: key
        for key, label in head.items()
        if key != "_id"
    }
    required = ("年份", "省份", "类型", "专业", "计划数", "最高分", "最低分", "平均分", "最低分位次")
    if not all(label in label_to_key for label in required):
        return []

    def cell(row: dict, label: str) -> str:
        return str(row.get(label_to_key[label], "") or "").strip()

    records: list[AdmissionScoreRecord] = []
    for row in data.get("list") or []:
        if cell(row, "省份") != config.province or cell(row, "类型") != batch:
            continue
        major_name = cell(row, "专业")
        min_score = _integer(cell(row, "最低分"))
        if not major_name or min_score is None:
            continue
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=int(cell(row, "年份") or provenance.year),
                batch=batch,
                subject_type="unified",
                university_code=target.university_code,
                university_name=target.name,
                major_group_code=major_name,
                major_group_name=major_name,
                min_score=min_score,
                max_score=_integer(cell(row, "最高分")),
                avg_score=_integer(cell(row, "平均分")),
                min_rank=_integer(cell(row, "最低分位次")),
                provenance=provenance,
            )
        )
    return records


def parse_zjnu_admission_score_json(
    path: Path,
    *,
    provenance: Provenance,
    config: PipelineConfig,
    target_university_code: str,
) -> list[AdmissionScoreRecord]:
    """浙江师范大学"历年分数"栏目是纯前端JS单页应用（`lqcx.zjnu.edu.cn/zsdata/
    lqxx/#/lnfs`），此前判定"httpx拿不到数据需要浏览器渲染"——这只说对了一半：
    确实不能直接抓这个URL（`#`后面是前端路由不是真实请求），但页面背后调用的
    是`POST /lqxx/s/api/front/lqxx/getList`这个JSON接口，已用真实响应验证
    httpx直接POST同样的body（`{type:"lnfs",sf,nf,zslb,klmc,xqmc}`）能拿到和
    浏览器里一模一样的数据，不需要Playwright；`Access-Control-Allow-Origin`
    锁定同源只影响浏览器端`fetch`，不影响服务端HTTP客户端。字段是拼音缩写
    （zymc=专业名称/lqrs=录取人数/zgf=最高分/pjf=平均分/zdf=最低分/zdfwc=
    最低位次），`lqrs`这里确实是"实际录取人数"（不是招生计划数），跟
    `AdmissionScoreRecord.enrollment_count`语义一致，可以直接填。
    """
    by_code = {target.university_code: target for target in config.target_universities}
    target = by_code.get(target_university_code)
    if target is None:
        raise ValueError(f"target_university_code {target_university_code!r} not in whitelist")

    payload = json.loads(path.read_text(encoding="utf-8"))
    records: list[AdmissionScoreRecord] = []
    for row in payload.get("list") or []:
        if row.get("sf") != config.province:
            continue
        major_name = str(row.get("zymc") or "").strip()
        min_score = _integer(str(row.get("zdf") or ""))
        if not major_name or min_score is None:
            continue
        # 艺术类同一专业名称常按主项拆成多条独立招生线（真实数据："音乐学（师范）"
        # 器乐类/声乐类分数线不同但专业名称字面完全相同），只用major_name当
        # natural_key会撞车被判定为重复记录而误拒——跟宁波大学"音乐学（师范）"
        # 器乐/声乐主项在AdmissionPlanRecord.restrictions里遇到的是同一类问题，
        # 这里没有restrictions字段，改成把类别并进major_group_name本身
        category = str(row.get("zslb") or "").strip()
        if category and category != "普通类":
            major_name = f"{major_name}（{category}）"
        records.append(
            AdmissionScoreRecord(
                province=config.province,
                year=int(row.get("nf") or provenance.year),
                batch=str(row.get("pcmc") or "").strip() or "本科批",
                subject_type="unified",
                university_code=target.university_code,
                university_name=target.name,
                major_group_code=major_name,
                major_group_name=major_name,
                min_score=min_score,
                max_score=_integer(str(row.get("zgf") or "")),
                avg_score=_integer(str(row.get("pjf") or "")),
                enrollment_count=_integer(str(row.get("lqrs") or "")),
                min_rank=_integer(str(row.get("zdfwc") or "")),
                provenance=provenance,
            )
        )
    return records
